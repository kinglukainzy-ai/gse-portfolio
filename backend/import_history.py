"""
One-time import: reads a GSE daily market data file (Excel or CSV) and
backfills the price_snapshots table. Expected GSE Market Reports format:

  Column A: Daily Date (DD/MM/YYYY text)
  Column B: Share Code (symbol, may have ** markers for suspended stocks)
  Column H: Closing Price - VWAP (GH₵)
  Column I: Price Change (GH₵)
  Column L: Total Shares Traded

Usage:
  python import_history.py "Daily Shares  ETFs 2023.csv"
  python import_history.py data.xlsx --dry-run
"""
import argparse
import csv
import logging

import db

logging.basicConfig(level=logging.INFO)
log = logging.getLogger("import-history")


def parse_date(raw: str) -> str | None:
    raw = str(raw).strip()
    parts = raw.split("/")
    if len(parts) == 3:
        dd, mm, yyyy = parts
        return f"{yyyy}-{mm.zfill(2)}-{dd.zfill(2)}"
    if len(raw) == 10 and raw[4] == "-":
        return raw
    return None


def parse_number(raw, default=0):
    if raw is None or raw == "":
        return default
    raw = str(raw).replace(",", "")
    try:
        return float(raw)
    except ValueError:
        return default


def iter_rows_csv(path: str):
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)
        for row in reader:
            if len(row) >= 12:
                yield row


def iter_rows_xlsx(path: str):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        yield row
    wb.close()


def import_file(path: str, dry_run: bool = False):
    if path.lower().endswith(".csv"):
        rows = iter_rows_csv(path)
    else:
        rows = iter_rows_xlsx(path)

    saved = 0
    skipped = 0

    for row in rows:
        raw_date = row[0]
        raw_symbol = row[1]
        raw_price = row[7]
        raw_change = row[8]
        raw_volume = row[11]

        if not raw_date or not raw_symbol or not raw_price:
            skipped += 1
            continue

        date = parse_date(str(raw_date))
        if date is None:
            log.warning("Skipping unparseable date: %s", raw_date)
            skipped += 1
            continue

        symbol = str(raw_symbol).strip().strip("*").strip().upper()
        if not symbol or not symbol.replace(" ", "").isalnum():
            skipped += 1
            continue

        price = parse_number(raw_price)
        change = parse_number(raw_change)
        volume = int(parse_number(raw_volume))

        if price <= 0:
            skipped += 1
            continue

        if dry_run:
            log.info("[DRY RUN] %s %s price=%.4f change=%.4f vol=%d", date, symbol, price, change, volume)
        else:
            db.save_snapshot(symbol, price, snap_date=date, change=change, volume=volume)
        saved += 1

    log.info("%s: %d rows imported, %d skipped.", "DRY RUN" if dry_run else "DONE", saved, skipped)


def main():
    parser = argparse.ArgumentParser(description="Import GSE historical data into snapshots DB")
    parser.add_argument("file", help="Path to the .xlsx or .csv file")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing to DB")
    args = parser.parse_args()

    db.init_db()
    import_file(args.file, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
